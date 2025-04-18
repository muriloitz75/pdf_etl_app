import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def load_to_excel(df, excel_path):
    """Exporta DataFrame para arquivo Excel com formatação básica"""
    # Criar diretório de saída se não existir
    output_dir = os.path.dirname(excel_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Exportar para Excel
    df.to_excel(excel_path, index=False, sheet_name='Dados')

    # Aplicar formatação
    wb = load_workbook(excel_path)
    ws = wb['Dados']

    # Definir estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Formatar cabeçalho
    for col_num, _ in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Formatar células de dados
    for row in range(2, len(df) + 2):  # +2 porque Excel é 1-indexado e temos cabeçalho
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            # Alinhar números à direita
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

            # Formatar datas
            column_name = df.columns[col-1].lower()
            if any(date_term in column_name for date_term in ['data', 'dt', 'date', 'emissão', 'emissao']):
                # Verificar se o valor da célula é uma data no formato DD/MM/YYYY
                if cell.value and re.match(r'\d{2}/\d{2}/\d{4}', str(cell.value)):
                    # Formatar como texto para garantir que seja exibido corretamente
                    cell.number_format = '@'
                    # Garantir que o valor seja uma string
                    cell.value = str(cell.value)

            # Formatar competência
            if 'competência' in column_name:
                # Verificar se o valor da célula é uma competência no formato MM/AAAA
                if cell.value and re.match(r'\d{2}/\d{4}', str(cell.value)):
                    # Formatar como texto para garantir que seja exibido corretamente
                    cell.number_format = '@'
                    # Garantir que o valor seja uma string
                    cell.value = str(cell.value)
                    # Centralizar o texto
                    cell.alignment = Alignment(horizontal='center')

            # Formatar valores monetários
            if any(value_term in column_name for value_term in ['valor', 'vlr', 'preço', 'preco', 'total']):
                cell.number_format = 'R$ #,##0.00'

    # Ajustar largura das colunas
    for col_num, _ in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        # Definir largura mínima e máxima
        max_length = 0
        for row_num in range(1, len(df) + 2):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
            max_length = max(max_length, len(cell_value))
        adjusted_width = min(max(max_length + 2, 10), 50)  # Entre 10 e 50 caracteres
        ws.column_dimensions[col_letter].width = adjusted_width

    # Congelar painel no cabeçalho
    ws.freeze_panes = 'A2'

    # Criar aba de resumo
    create_summary_sheet(wb, df)

    # Salvar o arquivo formatado
    wb.save(excel_path)

    return excel_path


def create_summary_sheet(wb, df):
    """Cria uma aba de resumo com informações agrupadas por competência"""
    # Verificar se a coluna de competência existe
    if 'competência' not in df.columns:
        return

    # Criar nova aba de resumo
    if 'Resumo' in wb.sheetnames:
        # Se a aba já existir, removê-la para recriar
        wb.remove(wb['Resumo'])

    ws_resumo = wb.create_sheet('Resumo', 0)  # Criar como primeira aba

    # Definir estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Criar cabeçalho da tabela de resumo
    headers = ['COMPETÊNCIA:', 'SITUAÇÃO:', 'RESUMO DA COMPETÊNCIA:', '', '', '', '', '', '']
    for col_num, header in enumerate(headers, 1):
        cell = ws_resumo.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Mesclar células para o cabeçalho "RESUMO DA COMPETÊNCIA:"
    ws_resumo.merge_cells('C1:I1')

    # Criar subcabeçalhos para o resumo da competência
    subcabeçalhos = ['NFs EMITIDAS:', 'NFs CANCELADAS:', 'NFs VÁLIDAS:', 'BC TRIBUTÁVEL:', 'ISS PRÓPRIO:', 'ISS RETIDO:']
    for col_num, subcabeçalho in enumerate(subcabeçalhos, 3):
        cell = ws_resumo.cell(row=2, column=col_num)
        cell.value = subcabeçalho
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Agrupar dados por competência
    competencias = df['competência'].unique()

    # Para cada competência, calcular os valores resumidos
    row_num = 3  # Começar na linha 3 (após os cabeçalhos)

    for competencia in sorted(competencias):
        if not competencia:  # Ignorar competências vazias
            continue

        # Filtrar dados da competência atual
        comp_data = df[df['competência'] == competencia].copy()

        # Remover registros duplicados para evitar contagem dupla
        comp_data = comp_data.drop_duplicates()

        # Contar NFs emitidas e canceladas
        nfs_emitidas = len(comp_data)
        nfs_canceladas = 0

        # Determinar a situação da competência
        situacao = 'EM ABERTO'  # Valor padrão

        if 'situação' in df.columns:
            # Identificar notas canceladas
            notas_canceladas = comp_data[comp_data['situação'].str.contains('CANCELAD', case=False, na=False)]
            nfs_canceladas = len(notas_canceladas)

            # Se todas as notas estiverem canceladas, a situação é CANCELADA
            if nfs_canceladas == nfs_emitidas and nfs_emitidas > 0:
                situacao = 'CANCELADA'
            else:
                # Caso contrário, pegar a situação mais comum entre as notas não canceladas
                notas_ativas = comp_data[~comp_data['situação'].str.contains('CANCELAD', case=False, na=False)]
                if len(notas_ativas) > 0:
                    # Contar ocorrências de cada situação
                    situacoes_count = notas_ativas['situação'].value_counts()
                    if len(situacoes_count) > 0:
                        # Pegar a situação mais comum
                        situacao = situacoes_count.index[0]

        # Filtrar apenas notas válidas (não canceladas) para cálculos financeiros
        notas_validas = comp_data
        if 'situação' in comp_data.columns:
            # Considerar tanto notas ESCRITURADAS quanto QUITADAS como válidas
            notas_validas = comp_data[~comp_data['situação'].str.contains('CANCELAD', case=False, na=False)]

        # Calcular base de cálculo tributável
        bc_tributavel = 0
        if 'base de cálculo' in notas_validas.columns and len(notas_validas) > 0:
            try:
                # Converter para numérico, ignorando erros
                base_calc_values = pd.to_numeric(
                    notas_validas['base de cálculo'].astype(str).str.replace(',', '.').str.replace('R$', '').str.strip(),
                    errors='coerce'
                )
                # Substituir valores NaN por 0
                base_calc_values = base_calc_values.fillna(0)
                # Remover valores extremos (opcional - descomente se necessário)
                # q_low = base_calc_values.quantile(0.01)
                # q_high = base_calc_values.quantile(0.99)
                # base_calc_values = base_calc_values[(base_calc_values >= q_low) & (base_calc_values <= q_high)]
                bc_tributavel = base_calc_values.sum()
            except Exception as e:
                print(f"Erro ao calcular base de cálculo para competência {competencia}: {str(e)}")

        # Calcular ISS próprio
        iss_proprio = 0
        if 'iss próprio' in notas_validas.columns and len(notas_validas) > 0:
            try:
                # Converter para numérico, ignorando erros
                iss_proprio_values = pd.to_numeric(
                    notas_validas['iss próprio'].astype(str).str.replace(',', '.').str.replace('R$', '').str.strip(),
                    errors='coerce'
                )
                # Substituir valores NaN por 0
                iss_proprio_values = iss_proprio_values.fillna(0)
                iss_proprio = iss_proprio_values.sum()
            except Exception as e:
                print(f"Erro ao calcular ISS próprio para competência {competencia}: {str(e)}")

        # Calcular ISS retido
        iss_retido = 0
        if 'iss retido' in notas_validas.columns and len(notas_validas) > 0:
            try:
                # Converter para numérico, ignorando erros
                iss_retido_values = pd.to_numeric(
                    notas_validas['iss retido'].astype(str).str.replace(',', '.').str.replace('R$', '').str.strip(),
                    errors='coerce'
                )
                # Substituir valores NaN por 0
                iss_retido_values = iss_retido_values.fillna(0)
                iss_retido = iss_retido_values.sum()
            except Exception as e:
                print(f"Erro ao calcular ISS retido para competência {competencia}: {str(e)}")

        # Calcular NFs válidas (emitidas - canceladas)
        nfs_validas = nfs_emitidas - nfs_canceladas

        # Adicionar linha de resumo
        ws_resumo.cell(row=row_num, column=1).value = competencia
        ws_resumo.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row_num, column=1).border = thin_border

        ws_resumo.cell(row=row_num, column=2).value = situacao
        ws_resumo.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row_num, column=2).border = thin_border

        ws_resumo.cell(row=row_num, column=3).value = nfs_emitidas
        ws_resumo.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row_num, column=3).border = thin_border

        ws_resumo.cell(row=row_num, column=4).value = nfs_canceladas
        ws_resumo.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row_num, column=4).border = thin_border

        ws_resumo.cell(row=row_num, column=5).value = nfs_validas
        ws_resumo.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row_num, column=5).border = thin_border

        ws_resumo.cell(row=row_num, column=6).value = bc_tributavel
        ws_resumo.cell(row=row_num, column=6).number_format = 'R$ #,##0.00'
        ws_resumo.cell(row=row_num, column=6).alignment = Alignment(horizontal='right')
        ws_resumo.cell(row=row_num, column=6).border = thin_border

        ws_resumo.cell(row=row_num, column=7).value = iss_proprio
        ws_resumo.cell(row=row_num, column=7).number_format = 'R$ #,##0.00'
        ws_resumo.cell(row=row_num, column=7).alignment = Alignment(horizontal='right')
        ws_resumo.cell(row=row_num, column=7).border = thin_border

        ws_resumo.cell(row=row_num, column=8).value = iss_retido
        ws_resumo.cell(row=row_num, column=8).number_format = 'R$ #,##0.00'
        ws_resumo.cell(row=row_num, column=8).alignment = Alignment(horizontal='right')
        ws_resumo.cell(row=row_num, column=8).border = thin_border

        row_num += 1

    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 15  # COMPETÊNCIA
    ws_resumo.column_dimensions['B'].width = 15  # SITUAÇÃO
    ws_resumo.column_dimensions['C'].width = 15  # NFs EMITIDAS
    ws_resumo.column_dimensions['D'].width = 15  # NFs CANCELADAS
    ws_resumo.column_dimensions['E'].width = 15  # NFs VÁLIDAS
    ws_resumo.column_dimensions['F'].width = 20  # BC TRIBUTÁVEL
    ws_resumo.column_dimensions['G'].width = 15  # ISS PRÓPRIO
    ws_resumo.column_dimensions['H'].width = 15  # ISS RETIDO

    # Congelar painel no cabeçalho
    ws_resumo.freeze_panes = 'A3'
